import csv
import io
import ipaddress
import json
import math
import socket
from pathlib import Path
from urllib.parse import urlparse
from xml.etree import ElementTree

from django.conf import settings
from rest_framework import serializers


ALLOWED_FILE_TYPES = {"pdf", "docx", "xlsx", "csv", "xml", "json"}


def validate_custom_text(value):
    value = (value or "").strip()
    if not value:
        raise serializers.ValidationError("Content cannot be empty.")
    if len(value.split()) > settings.KNOWLEDGE_MAX_TEXT_WORDS:
        raise serializers.ValidationError(
            f"Content cannot exceed {settings.KNOWLEDGE_MAX_TEXT_WORDS} words."
        )
    return value


def validate_public_url(value):
    parsed = urlparse(value)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise serializers.ValidationError("Enter a valid HTTP or HTTPS URL.")
    if parsed.username or parsed.password:
        raise serializers.ValidationError("URLs containing credentials are not allowed.")
    try:
        addresses = {
            item[4][0]
            for item in socket.getaddrinfo(
                parsed.hostname,
                parsed.port or (443 if parsed.scheme == "https" else 80),
                type=socket.SOCK_STREAM,
            )
        }
    except socket.gaierror as exc:
        raise serializers.ValidationError("The URL host could not be resolved.") from exc
    for address in addresses:
        if not ipaddress.ip_address(address).is_global:
            raise serializers.ValidationError(
                "Private, loopback, and reserved network URLs are not allowed."
            )
    return value


def validate_knowledge_file(uploaded_file):
    suffix = Path(uploaded_file.name).suffix.lower().lstrip(".")
    if suffix not in ALLOWED_FILE_TYPES:
        allowed = ", ".join(sorted(ALLOWED_FILE_TYPES))
        raise serializers.ValidationError(
            f"Unsupported file type. Allowed types: {allowed}."
        )
    max_bytes = settings.KNOWLEDGE_MAX_FILE_SIZE_MB * 1024 * 1024
    if uploaded_file.size > max_bytes:
        raise serializers.ValidationError(
            f"File cannot exceed {settings.KNOWLEDGE_MAX_FILE_SIZE_MB} MB."
        )
    if uploaded_file.size == 0:
        raise serializers.ValidationError("The uploaded file is empty.")
    try:
        validators = {
            "pdf": _validate_pdf,
            "docx": _validate_docx,
            "xlsx": _validate_xlsx,
            "csv": _validate_csv,
            "json": _validate_json,
            "xml": _validate_xml,
        }
        validators[suffix](uploaded_file)
    except serializers.ValidationError:
        raise
    except Exception as exc:
        raise serializers.ValidationError(
            f"The {suffix.upper()} file is invalid or corrupted."
        ) from exc
    finally:
        uploaded_file.seek(0)
    return uploaded_file


def _validate_pdf(file_obj):
    from pypdf import PdfReader

    if len(PdfReader(file_obj).pages) > settings.KNOWLEDGE_MAX_PDF_PAGES:
        raise serializers.ValidationError(
            f"PDF cannot exceed {settings.KNOWLEDGE_MAX_PDF_PAGES} pages."
        )


def _validate_docx(file_obj):
    from docx import Document

    document = Document(file_obj)
    words = sum(len(paragraph.text.split()) for paragraph in document.paragraphs)
    estimated_pages = math.ceil(words / max(settings.KNOWLEDGE_DOCX_WORDS_PER_PAGE, 1))
    if estimated_pages > settings.KNOWLEDGE_MAX_DOCX_PAGES:
        raise serializers.ValidationError(
            f"DOCX cannot exceed {settings.KNOWLEDGE_MAX_DOCX_PAGES} estimated pages."
        )


def _validate_xlsx(file_obj):
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        row_count = sum(sheet.max_row for sheet in workbook.worksheets)
    finally:
        workbook.close()
    if row_count > settings.KNOWLEDGE_MAX_SPREADSHEET_ROWS:
        raise serializers.ValidationError(
            f"XLSX cannot exceed {settings.KNOWLEDGE_MAX_SPREADSHEET_ROWS} total rows."
        )


def _read_text(file_obj):
    return file_obj.read().decode("utf-8-sig")


def _validate_csv(file_obj):
    row_count = sum(1 for _ in csv.reader(io.StringIO(_read_text(file_obj))))
    if row_count > settings.KNOWLEDGE_MAX_CSV_ROWS:
        raise serializers.ValidationError(
            f"CSV cannot exceed {settings.KNOWLEDGE_MAX_CSV_ROWS} rows."
        )


def _structured_item_count(value):
    if isinstance(value, dict):
        return len(value) + sum(_structured_item_count(item) for item in value.values())
    if isinstance(value, list):
        return len(value) + sum(_structured_item_count(item) for item in value)
    return 1


def _validate_json(file_obj):
    item_count = _structured_item_count(json.loads(_read_text(file_obj)))
    if item_count > settings.KNOWLEDGE_MAX_STRUCTURED_ITEMS:
        raise serializers.ValidationError(
            f"JSON cannot exceed {settings.KNOWLEDGE_MAX_STRUCTURED_ITEMS} items."
        )


def _validate_xml(file_obj):
    item_count = sum(1 for _ in ElementTree.parse(file_obj).iter())
    if item_count > settings.KNOWLEDGE_MAX_STRUCTURED_ITEMS:
        raise serializers.ValidationError(
            f"XML cannot exceed {settings.KNOWLEDGE_MAX_STRUCTURED_ITEMS} elements."
        )
